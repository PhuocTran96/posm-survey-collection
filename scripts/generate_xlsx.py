#!/usr/bin/env python3
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_stores_xlsx(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores"
    
    headers = ['Store ID', 'Store Code', 'Store Name', 'Channel', 'HC', 'Region', 'Province', 'MCP', 'TDL', 'TDS', 'Status', 'Created At']
    
    header_fill = PatternFill(start_color="0B5CAD", end_color="0B5CAD", fill_type="solid")
    header_font = Font(name='Times New Roman', color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, store in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=store.get('Store ID', ''))
        ws.cell(row=row_idx, column=2, value=store.get('Store Code', ''))
        ws.cell(row=row_idx, column=3, value=store.get('Store Name', ''))
        ws.cell(row=row_idx, column=4, value=store.get('Channel', ''))
        ws.cell(row=row_idx, column=5, value=store.get('HC', 0))
        ws.cell(row=row_idx, column=6, value=store.get('Region', ''))
        ws.cell(row=row_idx, column=7, value=store.get('Province', ''))
        ws.cell(row=row_idx, column=8, value=store.get('MCP', 'N'))
        ws.cell(row=row_idx, column=9, value=store.get('TDL', ''))
        ws.cell(row=row_idx, column=10, value=store.get('TDS', ''))
        ws.cell(row=row_idx, column=11, value=store.get('Status', 'Active'))
        ws.cell(row=row_idx, column=12, value=store.get('Created At', ''))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)

def create_users_xlsx(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    headers = ['User ID', 'Username', 'Login ID', 'Role', 'Leader', 'Status', 'Created At']
    
    header_fill = PatternFill(start_color="0B5CAD", end_color="0B5CAD", fill_type="solid")
    header_font = Font(name='Times New Roman', color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, user in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=user.get('User ID', ''))
        ws.cell(row=row_idx, column=2, value=user.get('Username', ''))
        ws.cell(row=row_idx, column=3, value=user.get('Login ID', ''))
        ws.cell(row=row_idx, column=4, value=user.get('Role', ''))
        ws.cell(row=row_idx, column=5, value=user.get('Leader', ''))
        ws.cell(row=row_idx, column=6, value=user.get('Status', 'Active'))
        ws.cell(row=row_idx, column=7, value=user.get('Created At', ''))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python generate_xlsx.py <type> <data_json> <output_path>")
        sys.exit(1)
    
    xlsx_type = sys.argv[1]
    data_json = sys.argv[2]
    output_path = sys.argv[3]
    
    data = json.loads(data_json)
    
    if xlsx_type == 'stores':
        create_stores_xlsx(data, output_path)
    elif xlsx_type == 'users':
        create_users_xlsx(data, output_path)
    else:
        print(f"Unknown type: {xlsx_type}")
        sys.exit(1)
    
    print(f"Created {output_path}")
